#!/usr/bin/env python3
"""Buscador de stock para el skill consulta-stock-rols.

Tres modos de uso:

1) Modo legacy (busqueda paso a paso):
   --ref / --ancho / --largo / --fabricacion / --solo-fabricacion / --alternativas-de
   Salida: JSON.

2) Modo completo (recomendado, una sola llamada por consulta):
   --consulta-completa --ref REF --ancho A --largo L [--formato markdown]
   Ejecuta validacion + stock + (si vacio) fabricacion + (si vacio) alternativas.

3) Modo administrativo:
   --listar-colecciones-disponibles, --validar-coleccion COL

Cuando --formato markdown, la salida es texto listo para pegar al cliente.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_EXCEL = SKILL_DIR / "data" / "Piezas terminadas.xlsx"
DEFAULT_EXCEL_FABRICACION = SKILL_DIR / "data" / "Piezas en fabricacion.xlsx"
COLECCIONES_JSON = SKILL_DIR / "references" / "colecciones.json"
ALTERNATIVAS_JSON = SKILL_DIR / "references" / "alternativas.json"

MIN_LONGITUD_NO_COMPROMETIDA_M = 0.5


# ---------------------------------------------------------------------------
# Modelos
# ---------------------------------------------------------------------------

@dataclass
class Pieza:
    descripcion: str
    lote: str
    ancho: float
    longitud_actual: float
    reservada_firme: float
    longitud_disponible: float
    reservada_temporal: float
    longitud_no_comprometida: float
    estado: str
    obs_revision: Optional[str]
    obs_venta: Optional[str]


@dataclass
class PiezaFabricacion:
    descripcion: str
    lote: str
    ancho: float
    longitud: float
    reservada_firme: float
    reservada_temporal: float
    longitud_no_comprometida: float
    estado: str
    estado_pieza: str
    fecha_planif_fin: Optional[str]
    fecha_entrega_requerida: Optional[str]
    fecha_retraso: Optional[str]
    fecha_disponibilidad: Optional[str]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _f(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _fecha_iso(v) -> Optional[str]:
    if v is None or v == "":
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
        except Exception:
            return str(v)
    return str(v)


def fmt_ancho(a: float) -> str:
    if a == int(a):
        return f"{int(a)} m"
    s = f"{a:.2f}".rstrip("0").rstrip(".")
    return s.replace(".", ",") + " m"


def fmt_longitud(l: float) -> str:
    return f"{l:.2f}".replace(".", ",") + " m"


def normalizar(texto: str) -> str:
    if not texto:
        return ""
    t = unicodedata.normalize("NFD", texto)
    t = t.encode("ascii", "ignore").decode("ascii")
    t = re.sub(r"\s+", " ", t).strip().upper()
    return t


def tokens(texto: str) -> list[str]:
    return [t for t in re.split(r"[^A-Z0-9]+", normalizar(texto)) if t]


def score_match(consulta: str, descripcion: str) -> float:
    nc = normalizar(consulta)
    nd = normalizar(descripcion)
    if not nc or not nd:
        return 0.0
    if nc == nd:
        return 1.0
    if nc in nd:
        return 0.9
    tc = tokens(consulta)
    td = set(tokens(descripcion))
    if not tc:
        return 0.0
    return sum(1 for t in tc if t in td) / len(tc)


# ---------------------------------------------------------------------------
# Carga de datos
# ---------------------------------------------------------------------------

def cargar_piezas(excel_path: Path) -> list[Pieza]:
    if not excel_path.exists():
        raise FileNotFoundError(f"No encuentro el Excel de stock en: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    piezas: list[Pieza] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[0] is None:
            continue
        piezas.append(Pieza(
            descripcion=str(row[0]).strip(),
            lote=str(row[1]).strip() if row[1] else "",
            ancho=_f(row[2]),
            longitud_actual=_f(row[3]),
            reservada_firme=_f(row[4]),
            longitud_disponible=_f(row[5]),
            reservada_temporal=_f(row[6]),
            longitud_no_comprometida=_f(row[7]),
            estado=str(row[8]).strip() if row[8] else "",
            obs_revision=(str(row[9]).strip() if row[9] else None),
            obs_venta=(str(row[10]).strip() if row[10] else None),
        ))
    return piezas


def cargar_piezas_fabricacion(excel_path: Path) -> list[PiezaFabricacion]:
    if not excel_path.exists():
        raise FileNotFoundError(f"No encuentro el Excel de fabricacion en: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    piezas: list[PiezaFabricacion] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or row[1] is None:
            continue
        fecha_planif = _fecha_iso(row[8])
        fecha_retraso = _fecha_iso(row[11])
        piezas.append(PiezaFabricacion(
            estado=str(row[0]).strip() if row[0] else "",
            descripcion=str(row[1]).strip(),
            lote=str(row[2]).strip() if row[2] else "",
            ancho=_f(row[3]),
            longitud=_f(row[4]),
            reservada_firme=_f(row[5]),
            reservada_temporal=_f(row[6]),
            longitud_no_comprometida=_f(row[7]),
            fecha_planif_fin=fecha_planif,
            estado_pieza=str(row[9]).strip() if row[9] else "",
            fecha_entrega_requerida=_fecha_iso(row[10]),
            fecha_retraso=fecha_retraso,
            fecha_disponibilidad=fecha_retraso or fecha_planif,
        ))
    return piezas


def cargar_colecciones() -> dict:
    if not COLECCIONES_JSON.exists():
        return {}
    with open(COLECCIONES_JSON, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("colecciones", {})


def cargar_alternativas() -> dict:
    if not ALTERNATIVAS_JSON.exists():
        return {}
    with open(ALTERNATIVAS_JSON, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Logica de busqueda
# ---------------------------------------------------------------------------

def detectar_coleccion(descripcion: str, colecciones: dict) -> Optional[str]:
    nd = normalizar(descripcion)
    candidatos = [c for c in colecciones if nd.startswith(normalizar(c))]
    if not candidatos:
        candidatos = [c for c in colecciones if normalizar(c) in nd]
    if not candidatos:
        return None
    return max(candidatos, key=lambda c: len(c))


def validar_medida_contra_coleccion(coleccion, ancho_pedido, largo_pedido, colecciones):
    spec = colecciones.get(coleccion)
    if not spec:
        return {"coleccion": coleccion, "conocida": False}
    anchos_rollo = sorted(spec["anchos_rollo"])
    max_w, max_l = spec["max_alfombra"]
    a_dir = next((a for a in anchos_rollo if a >= ancho_pedido), None) if ancho_pedido is not None else None
    a_rot = next((a for a in anchos_rollo if a >= largo_pedido), None) if largo_pedido is not None else None
    encaja_directo = (ancho_pedido is None) or (a_dir is not None)
    encaja_rotando = (
        largo_pedido is not None and a_rot is not None
        and (ancho_pedido is None or ancho_pedido <= max_l)
    )
    return {
        "coleccion": coleccion,
        "conocida": True,
        "anchos_rollo_disponibles": anchos_rollo,
        "max_alfombra": [max_w, max_l],
        "solo_alfombra": spec.get("solo_alfombra", False),
        "orientacion_importa": spec.get("orientacion_importa", True),
        "material": spec.get("material"),
        "exterior": spec.get("exterior", False),
        "nota": spec.get("nota"),
        "ancho_rollo_directo": a_dir,
        "ancho_rollo_rotando": a_rot,
        "encaja_directo": encaja_directo,
        "encaja_rotando": encaja_rotando,
    }


def buscar_alternativas(color: str, alternativas: dict) -> dict:
    if not alternativas:
        return {"found": False, "razon": "alternativas.json vacio o no existe"}
    nc = normalizar(color)
    for k, v in alternativas.items():
        if normalizar(k) == nc:
            return {"found": True, "color": k, **v}
    candidatos = [k for k in alternativas if nc in normalizar(k) or normalizar(k) in nc]
    if len(candidatos) == 1:
        k = candidatos[0]
        return {"found": True, "color": k, "match_tipo": "fuzzy", **alternativas[k]}
    if len(candidatos) > 1:
        return {"found": False, "razon": "ambiguo", "posibles": candidatos}
    return {"found": False, "razon": "no_encontrado"}


def buscar(piezas, ref=None, ancho_min=None, largo_min=None, estados=None,
           umbral_match=0.5, limite=50, incluir_retales=False):
    resultados = []
    for p in piezas:
        if not incluir_retales and p.longitud_no_comprometida < MIN_LONGITUD_NO_COMPROMETIDA_M:
            continue
        score = 1.0
        if ref:
            score = score_match(ref, p.descripcion)
            if score < umbral_match:
                continue
        if ancho_min is not None and p.ancho < ancho_min:
            continue
        if largo_min is not None and p.longitud_no_comprometida < largo_min:
            continue
        if estados and p.estado not in estados:
            continue
        resultados.append((score, p))
    resultados.sort(key=lambda x: (-x[0], -x[1].longitud_no_comprometida))
    return [{**asdict(p), "match_score": round(s, 2)} for s, p in resultados[:limite]]


def buscar_fabricacion(piezas_fab, ref=None, ancho_min=None, largo_min=None,
                       umbral_match=0.5, limite=50):
    cand_match = []
    cand_match_ancho = []
    for p in piezas_fab:
        score = 1.0
        if ref:
            score = score_match(ref, p.descripcion)
            if score < umbral_match:
                continue
        cand_match.append((score, p))
        if ancho_min is not None and p.ancho < ancho_min:
            continue
        cand_match_ancho.append((score, p))
    con_libre = [(s, p) for s, p in cand_match_ancho if p.longitud_no_comprometida > 0]
    if largo_min is not None:
        con_libre = [(s, p) for s, p in con_libre if p.longitud_no_comprometida >= largo_min]
    con_libre.sort(key=lambda it: (-it[0], -it[1].longitud_no_comprometida,
                                    it[1].fecha_disponibilidad or "9999-12-31"))
    piezas_out = [{**asdict(p), "match_score": round(s, 2)} for s, p in con_libre[:limite]]
    return {
        "piezas": piezas_out,
        "candidatas_con_match": len(cand_match),
        "candidatas_con_match_y_ancho": len(cand_match_ancho),
        "todas_comprometidas": bool(cand_match_ancho) and not con_libre,
    }


# ---------------------------------------------------------------------------
# Flujo completo en una pasada (--consulta-completa)
# ---------------------------------------------------------------------------

def consulta_completa(ref, ancho, largo, piezas_stock, piezas_fab,
                      colecciones, alternativas, limite=10, umbral_match=0.85):
    """Ejecuta el flujo: validacion -> stock -> (si vacio) fabricacion ->
    (si vacio) alternativas. Devuelve un dict consolidado.
    """
    out = {
        "consulta": {"ref": ref, "ancho": ancho, "largo": largo},
        "validacion": None,
        "stock": [],
        "fabricacion": None,
        "alternativas": None,
    }

    # 1) Validacion de coleccion (siempre que sepamos la coleccion, aunque
    # falte alguna medida — basta con un ancho excesivo para invalidar).
    coleccion = detectar_coleccion(ref, colecciones) if ref else None
    if coleccion:
        out["validacion"] = validar_medida_contra_coleccion(
            coleccion, ancho, largo, colecciones
        )

    # 2) Stock terminado
    out["stock"] = buscar(
        piezas_stock, ref=ref, ancho_min=ancho, largo_min=largo,
        umbral_match=umbral_match, limite=limite,
    )

    # 3) Si no hay stock, fabricacion
    if not out["stock"]:
        out["fabricacion"] = buscar_fabricacion(
            piezas_fab, ref=ref, ancho_min=ancho, largo_min=largo,
            umbral_match=umbral_match, limite=limite,
        )

    # 4) Si no hay stock, buscamos alternativas SIEMPRE (aunque haya
    # fabricacion): el cliente puede preferir algo en stock antes que esperar.
    if not out["stock"] and ref:
        alts = buscar_alternativas(ref, alternativas)
        out["alternativas"] = alts
        # Para cada alternativa propuesta, buscar piezas en stock que cumplan
        # las medidas pedidas. Esto evita que el agente tenga que iterar
        # manualmente; el frontend puede mostrar directamente lotes concretos.
        if alts and alts.get("found") and not alts.get("sin_alternativas"):
            for tier in ("tier_1", "tier_2"):
                for entry in alts.get(tier, []):
                    piezas_alt = buscar(
                        piezas_stock,
                        ref=entry["ref"],
                        ancho_min=ancho,
                        largo_min=largo,
                        umbral_match=umbral_match,
                        limite=5,
                    )
                    entry["piezas"] = piezas_alt
                    entry["tiene_stock"] = bool(piezas_alt)

    return out


# ---------------------------------------------------------------------------
# Renderizado Markdown
# ---------------------------------------------------------------------------

def _tabla_piezas(piezas: list[dict], modo: str) -> str:
    """modo='stock' o 'fabricacion'. Devuelve una tabla Markdown."""
    rows = [
        "| Lote | Ancho | Longitud no comprometida | Estado | Notas |",
        "|---|---|---|---|---|",
    ]
    for i, p in enumerate(piezas):
        lote = p["lote"]
        ancho_str = fmt_ancho(p["ancho"])
        libre = fmt_longitud(p["longitud_no_comprometida"])
        estado = p["estado"]
        if i == 0:
            lote = f"**{lote}**"
            libre = f"**{libre}**"
        notas = []
        if modo == "stock":
            rt = p.get("reservada_temporal", 0)
            if rt and rt > 0:
                notas.append(
                    f"⚠️ tiene {fmt_longitud(rt)} reservados temporalmente; validar antes de comprometer"
                )
            if p.get("obs_revision"):
                notas.append(f"obs. revisión: \"{p['obs_revision']}\"")
            if p.get("obs_venta"):
                notas.append(f"obs. venta: \"{p['obs_venta']}\"")
        elif modo == "fabricacion":
            if p.get("fecha_retraso"):
                notas.append(f"Retraso a {p['fecha_retraso']}")
            elif p.get("fecha_disponibilidad"):
                notas.append(f"Disp. ~{p['fecha_disponibilidad']}")
        nota_str = " · ".join(notas) if notas else "—"
        rows.append(
            f"| {lote} | {ancho_str} | {libre} | {estado} | {nota_str} |"
        )
    return "\n".join(rows)


def render_markdown(resultado: dict) -> str:
    ref = resultado["consulta"]["ref"] or ""
    ancho = resultado["consulta"]["ancho"]
    largo = resultado["consulta"]["largo"]

    medida_str = ""
    if ancho and largo:
        a = fmt_ancho(ancho).replace(" m", "")
        l = fmt_longitud(largo).replace(" m", "")
        medida_str = f" ({a} × {l} m)"

    out: list[str] = []

    # Caso 1: hay stock
    if resultado["stock"]:
        n = len(resultado["stock"])
        verbo = "Tenemos" if n > 1 else "Tenemos"
        out.append(f"{verbo} stock para **{ref}**{medida_str}:")
        out.append("")
        out.append(_tabla_piezas(resultado["stock"], modo="stock"))
        return "\n".join(out)

    fab = resultado.get("fabricacion") or {}

    # Caso 2: no stock, fabricacion con piezas libres
    if fab.get("piezas"):
        out.append(
            f"No tenemos stock terminado de **{ref}**{medida_str}, pero sí hay piezas en fabricación:"
        )
        out.append("")
        out.append(_tabla_piezas(fab["piezas"], modo="fabricacion"))
        return "\n".join(out)

    # Caso 3: no stock, fabricacion existe pero todas comprometidas
    if fab.get("todas_comprometidas"):
        out.append(
            f"En este momento no tenemos stock terminado de **{ref}**{medida_str} y "
            f"**toda la fabricación en curso de esa referencia ya está reservada**."
        )
        out.append("")
        out.append("¿Quieres que mire referencias similares?")
        _append_alternativas_para_agente(out, resultado.get("alternativas"))
        return "\n".join(out)

    # Caso 4: no stock, no fabricacion dada de alta
    out.append(
        f"No tenemos stock terminado de **{ref}**{medida_str} y todavía no hay fabricación "
        f"dada de alta para esa referencia."
    )
    out.append("")
    out.append(
        "Si te pones en contacto con nosotros, te confirmamos para cuándo entraría la próxima fabricación."
    )
    out.append("")
    out.append("¿Quieres que mire referencias similares ya fabricadas?")
    _append_alternativas_para_agente(out, resultado.get("alternativas"))
    return "\n".join(out)


def _append_alternativas_para_agente(out: list[str], alts: Optional[dict]) -> None:
    """Adjunta una seccion HTML-comment con sugerencias para que el agente las
    presente si el cliente acepta. NO se muestra al cliente directamente."""
    if not alts or not alts.get("found"):
        return
    if alts.get("sin_alternativas"):
        out.append("")
        out.append(
            "<!-- Para esta referencia el catálogo no tiene alternativas equivalentes "
            "(sin_alternativas: true). Si el cliente acepta opciones, ofrecer contacto comercial. -->"
        )
        return
    lines = ["<!-- Alternativas disponibles si el cliente acepta:"]
    if alts.get("tier_1"):
        lines.append("Tier I (misma familia, otro diseño):")
        for e in alts["tier_1"]:
            nota = f" — {e['nota']}" if e.get("nota") else ""
            lines.append(f"  - {e['ref']}{nota}")
    if alts.get("tier_2"):
        lines.append("Tier II (otra construcción / colorido similar):")
        for e in alts["tier_2"]:
            nota = f" — {e['nota']}" if e.get("nota") else ""
            lines.append(f"  - {e['ref']}{nota}")
    lines.append("-->")
    out.append("")
    out.append("\n".join(lines))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Buscador de stock Rols")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--excel-fabricacion", type=Path, default=DEFAULT_EXCEL_FABRICACION)
    parser.add_argument("--fabricacion", action="store_true")
    parser.add_argument("--solo-fabricacion", action="store_true")
    parser.add_argument("--consulta-completa", action="store_true",
                        help="Ejecuta validacion + stock + fabricacion + alternativas en una sola llamada")
    parser.add_argument("--formato", choices=["json", "markdown"], default="json",
                        help="Formato de salida (default json)")
    parser.add_argument("--ref")
    parser.add_argument("--ancho", type=float)
    parser.add_argument("--largo", type=float)
    parser.add_argument("--estados", nargs="+")
    parser.add_argument("--umbral", type=float, default=0.5)
    parser.add_argument("--limite", type=int, default=50)
    parser.add_argument("--incluir-retales", action="store_true")
    parser.add_argument("--listar-colecciones-disponibles", action="store_true")
    parser.add_argument("--validar-coleccion")
    parser.add_argument("--alternativas-de",
                        help="Devuelve alternativas Tier I / Tier II para el color")
    args = parser.parse_args()

    try:
        piezas = cargar_piezas(args.excel)
    except FileNotFoundError as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        return 2

    colecciones = cargar_colecciones()

    # --- Modos administrativos ---
    if args.listar_colecciones_disponibles:
        cols = sorted({detectar_coleccion(p.descripcion, colecciones) or p.descripcion.split()[0]
                       for p in piezas})
        print(json.dumps({"colecciones_en_stock": cols}, ensure_ascii=False, indent=2))
        return 0

    if args.validar_coleccion:
        col = args.validar_coleccion.upper()
        spec = colecciones.get(col)
        if spec is None:
            for k, v in colecciones.items():
                if normalizar(k) == normalizar(col):
                    spec = v; col = k; break
        print(json.dumps({"coleccion": col, "spec": spec}, ensure_ascii=False, indent=2))
        return 0

    if args.alternativas_de:
        alts = cargar_alternativas()
        print(json.dumps(buscar_alternativas(args.alternativas_de, alts),
                         ensure_ascii=False, indent=2))
        return 0

    # --- Modo consulta completa ---
    if args.consulta_completa:
        if not args.ref:
            print(json.dumps({"error": "--consulta-completa requiere --ref"},
                             ensure_ascii=False))
            return 2
        try:
            piezas_fab = cargar_piezas_fabricacion(args.excel_fabricacion)
        except FileNotFoundError:
            piezas_fab = []
        alternativas = cargar_alternativas()
        resultado = consulta_completa(
            args.ref, args.ancho, args.largo,
            piezas, piezas_fab, colecciones, alternativas,
            limite=args.limite,
        )
        if args.formato == "markdown":
            print(render_markdown(resultado))
        else:
            print(json.dumps(resultado, ensure_ascii=False, indent=2, default=str))
        return 0

    # --- Modo legacy (paso a paso) ---
    if args.solo_fabricacion:
        resultados = []
    else:
        resultados = buscar(
            piezas, ref=args.ref, ancho_min=args.ancho, largo_min=args.largo,
            estados=args.estados, umbral_match=args.umbral, limite=args.limite,
            incluir_retales=args.incluir_retales,
        )

    analisis_coleccion = None
    if args.ref and args.ancho and args.largo:
        coleccion = None
        if resultados:
            coleccion = detectar_coleccion(resultados[0]["descripcion"], colecciones)
        if not coleccion:
            coleccion = detectar_coleccion(args.ref, colecciones)
        if coleccion:
            analisis_coleccion = validar_medida_contra_coleccion(
                coleccion, args.ancho, args.largo, colecciones
            )

    fabricacion = None
    if args.solo_fabricacion or (args.fabricacion and not resultados):
        try:
            piezas_fab = cargar_piezas_fabricacion(args.excel_fabricacion)
            fabricacion = buscar_fabricacion(
                piezas_fab, ref=args.ref, ancho_min=args.ancho,
                largo_min=args.largo, umbral_match=args.umbral, limite=args.limite,
            )
        except FileNotFoundError as e:
            fabricacion = {"error": str(e)}

    salida = {
        "consulta": {
            "ref": args.ref, "ancho_min": args.ancho, "largo_min": args.largo,
            "estados": args.estados, "incluir_retales": args.incluir_retales,
            "fabricacion": args.fabricacion or args.solo_fabricacion,
        },
        "analisis_coleccion": analisis_coleccion,
        "n_resultados": len(resultados),
        "piezas": resultados,
        "fabricacion": fabricacion,
    }
    print(json.dumps(salida, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
